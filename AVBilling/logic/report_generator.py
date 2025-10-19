#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from __future__ import annotations

from collections import defaultdict
from typing import Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from utils.helpers import month_key
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.utils import ImageReader
from pathlib import Path
from config.defaults import app_paths


class ReportGenerator:
    def __init__(self, invoices: List[Dict]) -> None:
        self.invoices = invoices

    def monthly_summary(self) -> Dict[str, Dict[str, float]]:
        result: Dict[str, Dict[str, float]] = defaultdict(lambda: {"invoices": 0, "revenue": 0.0})
        for inv in self.invoices:
            mk = month_key(inv.get("date", ""))
            result[mk]["invoices"] += 1
            result[mk]["revenue"] += float(inv.get("grand_total", 0))
        return {k: {"invoices": v["invoices"], "revenue": round(v["revenue"], 2)} for k, v in result.items()}

    def sales_per_product(self) -> Dict[str, Dict[str, float]]:
        result: Dict[str, Dict[str, float]] = defaultdict(lambda: {"quantity": 0.0, "revenue": 0.0, "gst": 0.0})
        for inv in self.invoices:
            for it in inv.get("items", []):
                name = it.get("product_name", it.get("product_code", "Unknown"))
                qty = float(it.get("quantity", 0))
                result[name]["quantity"] += qty
                result[name]["revenue"] += float(it.get("line_total", it.get("total", 0)))
                result[name]["gst"] += float(it.get("gst_amount", it.get("gst", 0)))
        for k, v in result.items():
            v["quantity"] = round(v["quantity"], 2)
            v["revenue"] = round(v["revenue"], 2)
            v["gst"] = round(v["gst"], 2)
        return dict(result)

    def export_invoices_excel(self, path: str) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Invoices"
        headers = [
            "Date",
            "Invoice No",
            "Customer",
            "Subtotal",
            "Discount",
            "GST",
            "Total",
        ]
        ws.append(headers)
        for inv in self.invoices:
            ws.append(
                [
                    inv.get("date", ""),
                    inv.get("invoice_no", ""),
                    inv.get("customer_name", ""),
                    inv.get("subtotal", 0),
                    inv.get("discount_total", 0),
                    inv.get("gst_total", 0),
                    inv.get("grand_total", 0),
                ]
            )
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        wb.save(path)

    def export_sales_excel(self, path: str) -> None:
        sales = self.sales_per_product()
        wb = Workbook()
        ws = wb.active
        ws.title = "Sales"
        headers = ["Product", "Quantity", "GST", "Revenue"]
        ws.append(headers)
        for name, v in sales.items():
            ws.append([name, v["quantity"], v["gst"], v["revenue"]])
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        wb.save(path)

    def export_invoice_pdf(self, invoice: Dict, path: str) -> None:
        styles = getSampleStyleSheet()
        doc = SimpleDocTemplate(path, pagesize=A4, leftMargin=32, rightMargin=32, topMargin=32, bottomMargin=32)
        story = []
        # Template selection can alter style and columns
        template = (invoice.get("template") or "simple").lower()
        title_text = "AVBilling Invoice" if template == "simple" else ("Invoice - Detailed" if template == "detailed" else "Invoice (Compact)")
        title = Paragraph(title_text, styles["Title"])
        story.append(title)
        # Company/logo header
        company = (invoice.get("company") or "")
        if company:
            story.append(Paragraph(company, styles["Heading3"]))
        # Logo if available in settings
        try:
            logo_path = (app_paths()["settings"]).read_text(encoding="utf-8")  # dummy read to ensure path exists
            # Actually read JSON for logo
            import json
            settings = json.loads((app_paths()["settings"]).read_text(encoding="utf-8"))
            logo_file = settings.get("company", {}).get("logo_path", "")
            if logo_file and Path(logo_file).exists():
                story.append(Spacer(1, 6))
                story.append(Paragraph(" ", styles["Normal"]))
                img = ImageReader(logo_file)
                # Add with max width
                from reportlab.platypus import Image
                story.append(Image(logo_file, width=120, height=60))
                story.append(Spacer(1, 6))
        except Exception:
            pass
        meta = Paragraph(
            f"Invoice No: {invoice.get('invoice_no','')}<br/>Date: {invoice.get('date','')}<br/>Customer: {invoice.get('customer_name','')}",
            styles["Normal"],
        )
        story.append(meta)
        story.append(Spacer(1, 12))

        if template == "compact":
            data = [["Item", "Qty", "Total"]]
        else:
            data = [["Item", "Qty", "Rate", "GST%", "Total"]]
        for it in invoice.get("items", []):
            row = [str(it.get("product_name", ""))[:40], f"{it.get('quantity', 0)}"]
            if template != "compact":
                row.extend([f"{it.get('rate', 0)}", f"{it.get('gst', 0)}"])
            row.append(f"{float(it.get('line_total', it.get('total', 0))):.2f}")
            data.append(row)
        tbl = Table(data, repeatRows=1)
        tbl.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                    ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                ]
            )
        )
        story.append(tbl)
        story.append(Spacer(1, 12))

        totals = Paragraph(
            f"Subtotal: {invoice.get('subtotal', 0):.2f}<br/>"
            f"Discount: {invoice.get('discount_total', 0):.2f}<br/>"
            f"GST: {invoice.get('gst_total', 0):.2f}<br/>"
            f"Grand Total: {invoice.get('grand_total', 0):.2f}",
            styles["Normal"],
        )
        story.append(totals)
        doc.build(story)


